###########################################################################################
# Modules de l'IHM du logiciel RBRE version 1.0                                           #
#                                                                                         #
# Date : décembre 2021                                                                    #
# Auteurs : Aurélien Paumier                                                              #
# @courriel : aurelien.paumier@shom.fr                                                    #
# Copyright ; Shom 2021                                                                   #
#                                                                                         #
# Outils utilisés : Python 3.7.3, openpyxl, PyQt5 v5.14.2, Qt Designer v5.14.1            #
#                   QtWebENgine v5.14.0                                                   #
###########################################################################################
import sys
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog
import warnings
import openpyxl as xl


class WindowsScreen(QDialog):
    def __init__(self):
        super(WindowsScreen, self).__init__()
        self.directory = None
        self.file = None
        file_ui = r'.\config\Windows.ui'
        loadUi(file_ui, self)
        self.toolButton_open.clicked.connect(self.chooseFile)
        self.toolButton_save.clicked.connect(self.chooseDirectory)
        self.convertir.clicked.connect(self.Convert)
        self.quitter.clicked.connect(self.Quitter)

        self.show()

    def chooseFile(self):
        """Fonction slot de selection du fichier .xlsx et remplissage des champs de l'interface"""
        options = QFileDialog.Options()
        self.file = QFileDialog.getOpenFileName(self,
                                                "Sélectionnez un fichier .xlsx...",
                                                "C:\\",
                                                "XLSX (*.xlsx)",
                                                options=options)
        self.lineEdit_open.setText(self.file[0])

    def chooseDirectory(self):
        options = QFileDialog.Options()
        self.directory = QFileDialog.getExistingDirectory(self, options=options)
        print(self.directory)
        self.lineEdit_save.setText(self.directory)

    def Convert(self):
        rbr2tdb(self.file, self.directory)

    def Quitter(self):
        sys.exit(app.exec_())


def rbr2tdb(file_in, file_out):
    # Ouverture du fichier excel et récupération des données/metadata
    warnings.simplefilter("ignore")
    xlsx_file = xl.load_workbook(filename=file_in[0])
    warnings.simplefilter("default")
    metadata, rbr_data = xlsx_file.worksheets[0::2]
    serial = str(int(metadata['A11'].value))
    file_out = open(file_out+'/RBR_SN'+serial+".txt", "w")

    # Ecriture dans un fichier en sortie
    file_out.write("RBR n° "+serial+2*"\n")
    file_out.write("Date\tHeure\tPression en hectopascal\tTempérature en °C \n")
    for line in range(3, rbr_data.max_row + 1):
        text = str(rbr_data.cell(line, 1).value)
        new_text = text.replace('-', '/')
        file_out.write(new_text + "\t" + str(round(rbr_data.cell(line, 3).value * 100, 1)) + "\t" \
                       + str(round(rbr_data.cell(line, 2).value, 1)) + "\n")

    file_out.close()


app = QApplication(sys.argv)
welcome = WindowsScreen()
widget = QtWidgets.QStackedWidget()
widget.addWidget(welcome)
widget.setFixedHeight(460)
widget.setFixedWidth(600)
widget.show()

sys.exit(app.exec_())


